import requests
import openpyxl
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np


def get_embedding (text):
    api_key = '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'
    api_endpoint = "https://api.openai.com/v1/embeddings"
    headers = {
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',  # Modify based on the API requirements
    }
    payload_data = {
        'input': text,
        'model': 'text-embedding-ada-002',
    }

    response = requests.post (api_endpoint, headers=headers, json=payload_data)

    # Check the response status code
    if response.status_code == 200:
        # Request was successful
        data = response.json ()
        #print (data['data'][0]['embedding'])
        data_str = [str(value) for value in data['data'][0]['embedding']]
        return ','.join (data_str)

    else:
        # Request failed
        print (f"Error: {response.status_code} - {response.text}")
        return 'error'

def get_embeddingRetArray(text):
     api_key = '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'
     api_endpoint = "https://api.openai.com/v1/embeddings"
     headers = {
         'Authorization': f'Bearer {api_key}',
         'Content-Type': 'application/json',  # Modify based on the API requirements
     }
     payload_data = {
         'input': text,
         'model': 'text-embedding-ada-002',
     }

     response = requests.post (api_endpoint, headers=headers, json=payload_data)

     # Check the response status code
     if response.status_code == 200:
         # Request was successful
         data = response.json ()
         #print (data['data'][0]['embedding'])
         return data['data'][0]['embedding']

     else:
         # Request failed
         print (f"Error: {response.status_code} - {response.text}")
         return 'error'


def gen_embedding_excel (excel_file):
    #excel_file = "embedding_db.xlsx"

    wb_obj = openpyxl.load_workbook(excel_file)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(2, m_row + 1):
        cell_obj_read = sheet_obj.cell(row = i, column = 1)
        cell_obj_write = sheet_obj.cell(row = i, column = 2)
        print("Getting embedding for: ", cell_obj_read.value)
        cell_obj_write.value = get_embedding (cell_obj_read.value)

    wb_obj.save (excel_file)

def improve_promt (prompt, excel_file):
    #prompt = "Who is Hassan Djirdeh?"

    prompt_emb = get_embeddingRetArray (prompt)

    #excel_file = "embedding_db.xlsx"

    wb_obj = openpyxl.load_workbook(excel_file)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    prompt_addition = str()
    prv_cosine_sim_res = 0
    for i in range(2, m_row + 1):
        db_emb = sheet_obj.cell(row = i, column = 2).value
        cosine_sim_write = sheet_obj.cell(row = i, column = 3)
        db_emb_list = db_emb.split(',')
        db_emb_list_f = [float(value) for value in db_emb_list]
        print()
        print (db_emb_list_f)
        # Some Cosine formulae
        cosine_sim_res = 1 #cosine_similarity(np.array(prompt_emb), np.array(db_emb_list_f.split(','))
        if prv_cosine_sim_res < cosine_sim_res:
            prv_cosine_sim_res = cosine_sim_res
            prompt_addition = sheet_obj.cell(row = i, column = 1).value
    finalPrompt = f'Info: {prompt_addition} Question: {prompt} Answer:'
    print (finalPrompt)
    return finalPrompt

if __name__ == "__main__":
    improve_promt ("Who is Hassan Djirdeh?", "text_chunks.xlsx")