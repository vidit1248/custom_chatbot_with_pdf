from flask import Flask, request, jsonify
import tkinter as tk
from tkinter import filedialog
from flask_cors import CORS
import pdfplumber
from openpyxl import Workbook
import os
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
from excel_embedding_gen import *

DIR_PATH = os.path.dirname(os.path.realpath(__file__))
EXCEL_FILENAME = "text_chunks.xlsx"
EXCEL_FILESHEET = "TextChunks"


app = Flask(__name__)
CORS(app)  # Enable CORS for all origins

def select_files():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    print('Selected Files:')
    print(file_paths)
    return file_paths

def extract_text_from_pdfs(file_paths):
    all_text = ''
    print ("Got file path: ", file_paths)
    try:
        for file_path in file_paths:
            print('Extracting Text of the File ' + file_path)
            with pdfplumber.open(file_path) as pdf:
                pdf_text = ''
                for page in pdf.pages:
                    page_text = page.extract_text()
                    pdf_text += page_text.replace('\n', ' ').replace('\x0c', ' ').replace('•', '• ').replace('·', '· ')
                    pdf_text += ' '
                all_text+=pdf_text
        return all_text
    except Exception as e:
        return f"Error occurred: {str(e)}"

def divide_string_into_5(input_string, arrayLength):
    if len(input_string) < arrayLength:
        arrayLength = 1 #raise ValueError("Input string is too short to create an array of size 5")

    # Calculate the approximate length of each substring
    substring_length = len(input_string) // arrayLength
    remaining_chars = len(input_string) % arrayLength

    substrings = []
    start_index = 0

    # Divide the string into 5 parts
    for i in range(arrayLength):
        end_index = start_index + substring_length + (1 if i < remaining_chars else 0)
        substrings.append(input_string[start_index:end_index])
        start_index = end_index

    return substrings

def createExcelFromTextChunks(chunks):
    print("Generating Excel")
    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXCEL_FILESHEET

    # Write chunks into the Excel file
    for index, chunk in enumerate(chunks, start=1):
        print("Generating Data For Index:" + str(index))
        sheet.cell(row=index, column=1, value=chunk)
        sheet.cell(row=index, column=2, value=get_embedding(chunk))

    workbook.save(EXCEL_FILENAME )
    print("Text chunks exported to text_chunks.xlsx")

    #improve_promt ("database to store embedd", "text_chunks.xlsx")

@app.route('/get_file', methods=['GET'])
def get_file():
    text = extract_text_from_pdfs(select_files())
    createExcelFromTextChunks(divide_string_into_5(text, 10))
    file_path ='123456'
    if file_path:
        return jsonify({"file_path": file_path})
    else:
        return jsonify({"message": "No file selected."})

@app.route('/get_final_prompt', methods=['POST'])
def getEmbeddedString():
    promptString = request.get_json()
    print (promptString['content']) 
    #return jsonify({"finalPromptString": promptString['content']})

    finalString = improve_promt (promptString['content'], EXCEL_FILENAME)
    if finalString:
        return jsonify({"finalPromptString": finalString})
    else:
        return jsonify({"message": "No file selected."})

@app.route("/upload", methods=["POST"])
def upload():
    print ("Upload hit")
    # Get the list of files from webpage 
    files = request.files.getlist("File")
    print (files)
    # Iterate for each file in the files List, and Save them 
    for file in files:
        print (file.filename)
        file.save(file.filename)
        file_path_l = list()
        file_path_l.append(file.filename)
        text = extract_text_from_pdfs(file_path_l)
        createExcelFromTextChunks(divide_string_into_5(text, 10))
    return "{}"

if __name__ == '__main__':
    app.run(debug=True)  # Run the Flask app